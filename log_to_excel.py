"""
Module to parse training log to a Microsoft Excel file
"""
import argparse
import glob
import pathlib
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# pylint: disable=R0903
class LogToExcel:
    """
    Class that contains necessary function to do the parsing.
    In order for this to work, one must name the training logs with syntax as follow:\n
        training log_[algorithm name]_[samples per round]_[number_of_rounds].txt\n
    And also put all the log files in the "logs" folder or else it won't work.
    """

    def __init__(self, samples_per_round, number_of_rounds):
        self.__algo_list = []
        self.__rate_list = []
        self.__samples_per_round = samples_per_round
        self.__number_of_rounds = number_of_rounds

    def to_excel(self):
        """Write to a new Microsoft Excel sheet."""
        self.__parse_training_log()
        filename = "summary_" + str(self.__samples_per_round) + "_" \
                                + str(self.__number_of_rounds) + ".xlsx"
        self.__get_rate()
        if not self.__algo_list:
            print("There aren't any files with the specified parameters.")
        else:
            wbook = Workbook()
            wsheet = wbook.active
            wsheet.merge_cells(start_row=1, start_column=1,
                                    end_row=2, end_column=1)
            wsheet.cell(1, 1).alignment = Alignment(wrapText=True)
            wsheet.cell(1, 1).border = Border(diagonal=Side(border_style="thin"),
                                                diagonalDown=True)
            wsheet.cell(1, 1).font = Font(vertAlign="subscript")
            wsheet.cell(1, 1, " " * 24 + "Algorithm\nRound No.")
            for col in wsheet.iter_cols(min_row=3, min_col=1, max_row=self.__number_of_rounds+2):
                for i, cell in enumerate(col):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.value = i+1

            for i, elem in enumerate(self.__algo_list):
                wsheet.merge_cells(start_row=1, start_column=2*(i+1),
                                    end_row=1, end_column=2*(i+1)+1)
                wsheet.cell(1, 2*(i+1)).alignment = Alignment(horizontal="center",
                                                                    vertical="center")
                wsheet.cell(1, 2*(i+1), elem)
                wsheet.cell(2, 2*(i+1)).alignment = Alignment(horizontal="center",
                                                                    vertical="center")
                wsheet.cell(2, 2*(i+1), "before")
                wsheet.cell(2, 2*(i+1)+1).alignment = Alignment(horizontal="center",
                                                                    vertical="center")
                wsheet.cell(2, 2*(i+1)+1, "after")
                for j, row in enumerate(wsheet.iter_rows(min_row=3, min_col=2*(i+1),
                                        max_row=self.__number_of_rounds+2, max_col=2*(i+1)+1)):
                    for k, cell in enumerate(row):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.value = float(self.__rate_list[2*(i*self.__number_of_rounds+j)+k])
                        cell.number_format = "0.0000"

            for col in range(wsheet.min_column, wsheet.max_column + 1):
                wsheet.column_dimensions[get_column_letter(col)].bestFit = True

            wbook.save(filename)

    def __parse_training_log(self):
        """Parse training logs filename to corresponding params."""
        logs = list(glob.glob("logs/training log*.txt"))
        for elem in logs:
            temp = pathlib.Path(elem).stem.replace("training log_", "")
            params_list = temp.split("_")
            if self.__samples_per_round == int(params_list[1]) \
                        and self.__number_of_rounds == int(params_list[2]):
                self.__algo_list.append(params_list[0])

    def __get_rate(self):
        """Get the rate from the logs."""
        for elem in self.__algo_list:
            with open("logs/training log_" + elem + "_"
                                    + str(self.__samples_per_round) + "_"
                                    + str(self.__number_of_rounds) + ".txt",
                                    encoding="utf-8") as file:
                content = file.read()
            self.__rate_list.extend(re.findall(r"\d+\.\d+", content))

if __name__=="__main__":
    parser = argparse.ArgumentParser(description="Parse training log files to .xlsx")
    parser.add_argument("samples_per_round", type=int, help="samples per round")
    parser.add_argument("number_of_rounds", type=int, help="number of rounds")
    args = parser.parse_args()

    obj = LogToExcel(args.samples_per_round, args.number_of_rounds)
    obj.to_excel()
